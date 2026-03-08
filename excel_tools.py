"""
Excel操作ツール群 - openpyxlを使用してExcelファイルを操作する
"""
import os
import re
import json
from typing import Any, Optional
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string


# 現在開いているワークブックのキャッシュ
_workbook_cache: dict[str, Workbook] = {}

# 操作を許可するベースディレクトリ（起動時のカレントディレクトリ）
_BASE_DIR = os.path.realpath(os.getcwd())

_HEX_COLOR_RE = re.compile(r'^[0-9A-Fa-f]{6}$')
_VALID_ALIGN   = {"left", "center", "right"}


def _reraise_if_fatal(e: Exception) -> None:
    """KeyboardInterrupt / SystemExit は握りつぶさず再 raise する"""
    if isinstance(e, (KeyboardInterrupt, SystemExit)):
        raise


def _safe_path(file_path: str) -> str:
    """パストラバーサルを防ぐ。BASE_DIR 外のパスは拒否する。"""
    abs_path = os.path.realpath(os.path.join(_BASE_DIR, file_path))
    if not abs_path.startswith(_BASE_DIR + os.sep) and abs_path != _BASE_DIR:
        raise PermissionError(
            f"アクセス拒否: 作業ディレクトリ外のパスは操作できません: {abs_path}"
        )
    return abs_path


def _validate_hex_color(value: str, name: str) -> str:
    """HEXカラー文字列を検証して正規化する（例: '#FF0000' → 'FF0000'）"""
    normalized = value.lstrip("#").upper()
    if not _HEX_COLOR_RE.match(normalized):
        raise ValueError(f"{name} は6桁の16進数カラーコードで指定してください（例: FF0000）")
    return normalized


def _get_workbook(file_path: str, create_if_missing: bool = False) -> Workbook:
    abs_path = _safe_path(file_path)
    if abs_path in _workbook_cache:
        return _workbook_cache[abs_path]
    if os.path.exists(abs_path):
        wb = load_workbook(abs_path)
    elif create_if_missing:
        wb = Workbook()
        # デフォルトシート名を "Sheet1" に統一
        wb.active.title = "Sheet1"
    else:
        raise FileNotFoundError(f"ファイルが見つかりません: {abs_path}")
    _workbook_cache[abs_path] = wb
    return wb


def _cell_value(cell) -> Any:
    """セルの値を返す。日付型は文字列に変換"""
    import datetime
    if isinstance(cell.value, (datetime.datetime, datetime.date)):
        return str(cell.value)
    return cell.value


# ── ツール関数 ──────────────────────────────────────────────────────────────


def open_excel(file_path: str, create_if_missing: bool = False) -> dict:
    """Excelファイルを開く（またはキャッシュに読み込む）"""
    try:
        wb = _get_workbook(file_path, create_if_missing=create_if_missing)
        sheets = wb.sheetnames
        return {
            "success": True,
            "file_path": os.path.abspath(file_path),
            "sheets": sheets,
            "message": f"ファイルを開きました。シート: {', '.join(sheets)}"
        }
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def list_sheets(file_path: str) -> dict:
    """シート一覧を取得する"""
    try:
        wb = _get_workbook(file_path)
        return {"success": True, "sheets": wb.sheetnames}
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def read_sheet(file_path: str, sheet_name: Optional[str] = None,
               min_row: int = 1, max_row: Optional[int] = None,
               min_col: int = 1, max_col: Optional[int] = None) -> dict:
    """シートの内容を読み取る"""
    try:
        wb = _get_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        data = []
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row or ws.max_row,
            min_col=min_col, max_col=max_col or ws.max_column
        ):
            row_data = []
            for cell in row:
                row_data.append({
                    "address": cell.coordinate,
                    "value": _cell_value(cell)
                })
            data.append(row_data)

        # テーブル形式の文字列も生成
        table_lines = []
        for row in data:
            table_lines.append(" | ".join(
                str(c["value"]) if c["value"] is not None else ""
                for c in row
            ))
        table_str = "\n".join(table_lines)

        return {
            "success": True,
            "sheet": ws.title,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "data": data,
            "table": table_str
        }
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def read_cell(file_path: str, cell_address: str,
              sheet_name: Optional[str] = None) -> dict:
    """特定のセルの値を読み取る（例: A1, B2）"""
    try:
        wb = _get_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        cell = ws[cell_address.upper()]
        return {
            "success": True,
            "address": cell.coordinate,
            "value": _cell_value(cell),
            "data_type": cell.data_type,
            "sheet": ws.title
        }
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def write_cell(file_path: str, cell_address: str, value: Any,
               sheet_name: Optional[str] = None) -> dict:
    """特定のセルに値を書き込む"""
    try:
        wb = _get_workbook(file_path, create_if_missing=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        ws[cell_address.upper()] = value
        wb.save(os.path.abspath(file_path))
        return {
            "success": True,
            "message": f"{ws.title}!{cell_address.upper()} に '{value}' を書き込みました"
        }
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def write_range(file_path: str, start_cell: str, data: list[list],
                sheet_name: Optional[str] = None,
                create_if_missing: bool = True) -> dict:
    """開始セルから2次元配列データを書き込む（例: start_cell='A1', data=[[1,2],[3,4]]）"""
    try:
        wb = _get_workbook(file_path, create_if_missing=create_if_missing)
        ws = wb[sheet_name] if sheet_name else wb.active

        # 開始セルの行・列を取得
        start = ws[start_cell.upper()]
        start_row = start.row
        start_col = start.column

        written = 0
        for r_idx, row in enumerate(data):
            for c_idx, val in enumerate(row):
                ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
                written += 1

        wb.save(os.path.abspath(file_path))
        return {
            "success": True,
            "message": f"{written}個のセルにデータを書き込みました",
            "range": f"{start_cell.upper()}:{get_column_letter(start_col + max(len(r) for r in data) - 1)}{start_row + len(data) - 1}"
        }
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def apply_formula(file_path: str, cell_address: str, formula: str,
                  sheet_name: Optional[str] = None) -> dict:
    """セルに数式を設定する（例: formula='=SUM(A1:A10)'）"""
    try:
        wb = _get_workbook(file_path, create_if_missing=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        if not formula.startswith("="):
            formula = "=" + formula
        ws[cell_address.upper()] = formula
        wb.save(os.path.abspath(file_path))
        return {
            "success": True,
            "message": f"{ws.title}!{cell_address.upper()} に数式 '{formula}' を設定しました"
        }
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def create_sheet(file_path: str, sheet_name: str,
                 position: Optional[int] = None) -> dict:
    """新しいシートを作成する"""
    try:
        wb = _get_workbook(file_path, create_if_missing=True)
        if sheet_name in wb.sheetnames:
            return {"success": False, "error": f"シート '{sheet_name}' は既に存在します"}
        ws = wb.create_sheet(title=sheet_name, index=position)
        wb.save(os.path.abspath(file_path))
        return {
            "success": True,
            "message": f"シート '{sheet_name}' を作成しました",
            "sheets": wb.sheetnames
        }
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def delete_sheet(file_path: str, sheet_name: str) -> dict:
    """シートを削除する"""
    try:
        wb = _get_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            return {"success": False, "error": f"シート '{sheet_name}' が見つかりません"}
        if len(wb.sheetnames) == 1:
            return {"success": False, "error": "最後のシートは削除できません"}
        del wb[sheet_name]
        wb.save(os.path.abspath(file_path))
        return {
            "success": True,
            "message": f"シート '{sheet_name}' を削除しました",
            "sheets": wb.sheetnames
        }
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def format_cell(file_path: str, cell_address: str,
                bold: Optional[bool] = None,
                italic: Optional[bool] = None,
                font_size: Optional[int] = None,
                font_color: Optional[str] = None,
                bg_color: Optional[str] = None,
                horizontal_align: Optional[str] = None,
                number_format: Optional[str] = None,
                sheet_name: Optional[str] = None) -> dict:
    """セルの書式を設定する（色はHEX形式: 'FF0000' = 赤）"""
    try:
        wb = _get_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        cell = ws[cell_address.upper()]

        # フォント設定
        font_kwargs = {}
        if bold is not None:
            font_kwargs["bold"] = bold
        if italic is not None:
            font_kwargs["italic"] = italic
        if font_size is not None:
            font_kwargs["size"] = font_size
        if font_color is not None:
            font_kwargs["color"] = _validate_hex_color(font_color, "font_color")
        if font_kwargs:
            # 既存フォントを引き継ぎながら更新
            existing = cell.font
            cell.font = Font(
                bold=font_kwargs.get("bold", existing.bold),
                italic=font_kwargs.get("italic", existing.italic),
                size=font_kwargs.get("size", existing.size),
                color=font_kwargs.get("color", (existing.color.rgb if existing.color and existing.color.type == "rgb" else "000000"))
            )

        # 背景色
        if bg_color is not None:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=_validate_hex_color(bg_color, "bg_color")
            )

        # 水平配置
        if horizontal_align is not None:
            if horizontal_align not in _VALID_ALIGN:
                raise ValueError(f"horizontal_align は {_VALID_ALIGN} のいずれかを指定してください")
            cell.alignment = Alignment(horizontal=horizontal_align)

        # 数値書式
        if number_format is not None:
            cell.number_format = number_format

        wb.save(os.path.abspath(file_path))
        return {
            "success": True,
            "message": f"{cell_address.upper()} の書式を更新しました"
        }
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def set_column_width(file_path: str, column: str, width: float,
                     sheet_name: Optional[str] = None) -> dict:
    """列幅を設定する（column: 'A' や '1'）"""
    try:
        wb = _get_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        col_letter = column.upper() if not column.isdigit() else get_column_letter(int(column))
        ws.column_dimensions[col_letter].width = width
        wb.save(os.path.abspath(file_path))
        return {"success": True, "message": f"列 {col_letter} の幅を {width} に設定しました"}
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def save_excel(file_path: str, save_as: Optional[str] = None) -> dict:
    """ファイルを保存する（save_as を指定すると別名保存）"""
    try:
        wb = _get_workbook(file_path)
        target = os.path.abspath(save_as if save_as else file_path)
        wb.save(target)
        if save_as:
            # 新しいパスでもキャッシュ登録
            _workbook_cache[target] = wb
        return {"success": True, "message": f"'{target}' に保存しました"}
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


def get_sheet_info(file_path: str, sheet_name: Optional[str] = None) -> dict:
    """シートの基本情報（行数・列数・使用範囲）を返す"""
    try:
        wb = _get_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        return {
            "success": True,
            "sheet": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "dimensions": ws.dimensions,
            "all_sheets": wb.sheetnames
        }
    except Exception as e:
        _reraise_if_fatal(e)
        return {"success": False, "error": str(e)}


# ── ツール定義（Claude API用）──────────────────────────────────────────────

TOOLS = [
    {
        "name": "open_excel",
        "description": "Open an Excel file. If the file does not exist, set create_if_missing=true to create it.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file (.xlsx)"},
                "create_if_missing": {"type": "boolean", "description": "Create the file if it does not exist (default: false)"}
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "list_sheets",
        "description": "List all sheet names in an Excel file.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"}
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "read_sheet",
        "description": "Read the contents of a sheet. Optionally limit to a row/column range.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"},
                "sheet_name": {"type": "string", "description": "Sheet name (defaults to active sheet)"},
                "min_row": {"type": "integer", "description": "First row to read (default: 1)"},
                "max_row": {"type": "integer", "description": "Last row to read (default: last row)"},
                "min_col": {"type": "integer", "description": "First column to read (default: 1)"},
                "max_col": {"type": "integer", "description": "Last column to read (default: last column)"}
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "read_cell",
        "description": "Read the value of a specific cell (e.g. A1, B3).",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"},
                "cell_address": {"type": "string", "description": "Cell address, e.g. A1 or B3"},
                "sheet_name": {"type": "string", "description": "Sheet name (defaults to active sheet)"}
            },
            "required": ["file_path", "cell_address"]
        }
    },
    {
        "name": "write_cell",
        "description": "Write a value to a specific cell.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"},
                "cell_address": {"type": "string", "description": "Cell address, e.g. A1 or B3"},
                "value": {"description": "Value to write (string, number, or boolean)"},
                "sheet_name": {"type": "string", "description": "Sheet name (defaults to active sheet)"}
            },
            "required": ["file_path", "cell_address", "value"]
        }
    },
    {
        "name": "write_range",
        "description": "Write a 2D array of data starting from a given cell. Example: data=[[\"Name\",\"Age\"],[\"Alice\",30]]",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"},
                "start_cell": {"type": "string", "description": "Top-left cell to start writing, e.g. A1"},
                "data": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": "2D array (list of rows)"
                },
                "sheet_name": {"type": "string", "description": "Sheet name (defaults to active sheet)"}
            },
            "required": ["file_path", "start_cell", "data"]
        }
    },
    {
        "name": "apply_formula",
        "description": "Set a formula in a cell, e.g. '=SUM(A1:A10)' or '=AVERAGE(B2:B10)'.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"},
                "cell_address": {"type": "string", "description": "Target cell address, e.g. C1"},
                "formula": {"type": "string", "description": "Formula string starting with '=', e.g. =SUM(A1:A10)"},
                "sheet_name": {"type": "string", "description": "Sheet name (defaults to active sheet)"}
            },
            "required": ["file_path", "cell_address", "formula"]
        }
    },
    {
        "name": "create_sheet",
        "description": "Create a new sheet in the workbook.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"},
                "sheet_name": {"type": "string", "description": "Name of the new sheet"},
                "position": {"type": "integer", "description": "Insert position (0-based index, default: append at end)"}
            },
            "required": ["file_path", "sheet_name"]
        }
    },
    {
        "name": "delete_sheet",
        "description": "Delete a sheet from the workbook.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"},
                "sheet_name": {"type": "string", "description": "Name of the sheet to delete"}
            },
            "required": ["file_path", "sheet_name"]
        }
    },
    {
        "name": "format_cell",
        "description": "Apply formatting to a cell: bold, italic, font size, font color, background color, or alignment.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"},
                "cell_address": {"type": "string", "description": "Cell address, e.g. A1"},
                "bold": {"type": "boolean", "description": "Set bold"},
                "italic": {"type": "boolean", "description": "Set italic"},
                "font_size": {"type": "integer", "description": "Font size in pt"},
                "font_color": {"type": "string", "description": "Font color as hex string, e.g. FF0000 for red"},
                "bg_color": {"type": "string", "description": "Background fill color as hex string, e.g. FFFF00 for yellow"},
                "horizontal_align": {"type": "string", "enum": ["left", "center", "right"], "description": "Horizontal alignment"},
                "number_format": {"type": "string", "description": "Excel number format string, e.g. '#,##0' for integer with comma, '#,##0.00' for 2 decimal places, '0%' for percentage, '¥#,##0' for yen"},
                "sheet_name": {"type": "string", "description": "Sheet name (defaults to active sheet)"}
            },
            "required": ["file_path", "cell_address"]
        }
    },
    {
        "name": "set_column_width",
        "description": "Set the width of a column.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"},
                "column": {"type": "string", "description": "Column letter, e.g. A or B"},
                "width": {"type": "number", "description": "Column width value"},
                "sheet_name": {"type": "string", "description": "Sheet name (defaults to active sheet)"}
            },
            "required": ["file_path", "column", "width"]
        }
    },
    {
        "name": "save_excel",
        "description": "Save the Excel file. Optionally save to a new path.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"},
                "save_as": {"type": "string", "description": "New path to save a copy (omit to overwrite original)"}
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "get_sheet_info",
        "description": "Get basic info about a sheet: row count, column count, and used range.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the Excel file"},
                "sheet_name": {"type": "string", "description": "Sheet name (defaults to active sheet)"}
            },
            "required": ["file_path"]
        }
    }
]

# ツール名 → 関数のマッピング
TOOL_FUNCTIONS = {
    "open_excel": open_excel,
    "list_sheets": list_sheets,
    "read_sheet": read_sheet,
    "read_cell": read_cell,
    "write_cell": write_cell,
    "write_range": write_range,
    "apply_formula": apply_formula,
    "create_sheet": create_sheet,
    "delete_sheet": delete_sheet,
    "format_cell": format_cell,
    "set_column_width": set_column_width,
    "save_excel": save_excel,
    "get_sheet_info": get_sheet_info,
}


def execute_tool(tool_name: str, tool_input: dict) -> str:
    """ツールを実行してJSON文字列で結果を返す"""
    func = TOOL_FUNCTIONS.get(tool_name)
    if not func:
        return json.dumps({"success": False, "error": f"不明なツール: {tool_name}"}, ensure_ascii=False)
    result = func(**tool_input)
    return json.dumps(result, ensure_ascii=False, indent=2)
